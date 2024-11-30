import openpyxl
from xls2xlsx import XLS2XLSX


def get_dataset(filename: str):
    try:
        wb = openpyxl.load_workbook(filename)
    except:
        XLS2XLSX(filename).to_xlsx(filename + "x")
        wb = openpyxl.load_workbook(filename + "x")
    sheet = wb.active
    data = []

    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    students_data = []
    current_student_data = []
    for item in data:
        if item[0] == 'Текущая успеваемость учащегося':
            if current_student_data:
                students_data.append(current_student_data[:-2])
            current_student_data = [item]
        else:
            current_student_data.append(item)

    if current_student_data:
        students_data.append(current_student_data[:-2])

    dataset = []

    for student_data in students_data:
        student_dict = {}


        for item in student_data:
            if item[0] == 'ФИО учащегося:':
                student_dict['name'] = item[1]
            elif item[0] == 'Класс:':
                student_dict['class'] = item[1]
            elif item[0] == 'Кл. руководитель:':
                student_dict['teacher'] = item[1]
            elif item[0] == 'Период формирования успеваемости:':
                student_dict['period'] = item[1]
            elif item[0] == 'Дата формирования:':
                student_dict['date'] = item[1]

        student_dict['data'] = []


        for item in student_data[7:]:
            subject_dict = {}
            subject_dict['subj_name'] = item[0]

            subject_dict['marks'] = [mark for mark in item[1].split(',') if mark]

            subject_dict['average'] = float(item[2])
            student_dict['data'].append(subject_dict)

        if len(student_dict['data']) != 0:
            student_dict['average'] = sum(subject['average'] for subject in student_dict['data']) / len(student_dict['data'])
            dataset.append(student_dict)

    wb.close()

    return dataset
