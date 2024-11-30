import openpyxl
from xls2xlsx import XLS2XLSX
from app import db
from app.models import StudentRecord, Title
import uuid
import json


def get_dataset(filename: str):
    try:
        wb = openpyxl.load_workbook(filename)
    except Exception as e:
        print(f"Ошибка при загрузке файла: {e}")
        XLS2XLSX(filename).to_xlsx(filename + "x")
        wb = openpyxl.load_workbook(filename + "x")
    sheet = wb.active
    data = []

    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    students_data = []
    current_student_data = []
    for item in data:
        if item[0] == 'Текущая успеваемость учащегося':
            if current_student_data:
                students_data.append(current_student_data[:-2])
            current_student_data = [item]
        else:
            current_student_data.append(item)

    if current_student_data:
        students_data.append(current_student_data[:-2])

    dataset = []

    for student_data in students_data:
        student_dict = {}


        for item in student_data:
            if item[0] == 'ФИО учащегося:':
                student_dict['name'] = item[1]
            elif item[0] == 'Класс:':
                student_dict['class'] = item[1]
            elif item[0] == 'Кл. руководитель:':
                student_dict['teacher'] = item[1]
            elif item[0] == 'Период формирования успеваемости:':
                student_dict['period'] = item[1]
            elif item[0] == 'Дата формирования:':
                student_dict['date'] = item[1]

        student_dict['subjects'] = []


        for item in student_data[7:]:
            subject_dict = {}
            subject_dict['subj_name'] = item[0]
            try:
                subject_dict['marks'] = [mark for mark in item[1].split(',') if mark]
                subject_dict['average'] = float(item[2])
            except Exception as e:
                print(f"Ошибка при обработке предмета {item[0]}: {e}")
            student_dict['subjects'].append(subject_dict)

        dataset.append(student_dict)

    wb.close()

    print(f"Данные загружены для {len(dataset)} студентов.")
    return dataset


def calculate_overall_average(student_data):
    total_average = sum(subject['average'] for subject in student_data['subjects'])
    try:
        overall_average = total_average / len(student_data['subjects'])
        return overall_average
    except ZeroDivisionError:
        return 0


def calculate_needed_grades(student_data):
    needed_grades = []

    for subject in student_data['subjects']:
        current_average = subject['average']
        grades_needed = {"subject": subject['subj_name'], "fours_needed": 0, "fives_needed": 0}
        total_marks = len(subject['marks'])


        if current_average >= 4.51:
            needed_grades.append(grades_needed)
            continue


        max_iterations = 100


        if current_average >= 4.0:
            grades_needed["fives_needed"] = calculate_needed_grades_to_reach_target(current_average, total_marks, 4.51, 5, max_iterations)


        elif current_average < 4.0:
            grades_needed["fours_needed"] = calculate_needed_grades_to_reach_target(current_average, total_marks, 4.0, 4, max_iterations)
            current_average = (current_average * total_marks + grades_needed["fours_needed"] * 4) / (total_marks + grades_needed["fours_needed"])
            total_marks += grades_needed["fours_needed"]
            grades_needed["fives_needed"] = calculate_needed_grades_to_reach_target(current_average, total_marks, 4.51, 5, max_iterations)

        needed_grades.append(grades_needed)

    return needed_grades

def calculate_needed_grades_to_reach_target(current_average, total_marks, target_average, grade_value, max_iterations):
    needed_grades = 0
    while current_average < target_average and max_iterations > 0:
        needed_grades += 1
        current_average = (current_average * total_marks + grade_value) / (total_marks + 1)
        total_marks += 1
        max_iterations -= 1

    if max_iterations == 0:
        print(f"Внимание: Расчет необходимых оценок был остановлен из-за превышения лимита итераций.")

    return needed_grades



def assign_title(average_score):
    if 4.9 <= average_score <= 5.0:
        return "Великий Магистр Высшего Ранга (5+)"
    elif 4.5 <= average_score < 4.9:
        return "Магистр Всеведущего Разума (5)"
    elif 4.0 <= average_score < 4.5:
        return "Архимаг Мудрости (4+)"
    elif 3.7 <= average_score < 4.0:
        return "Великий Хранитель Знаний (4)"
    elif 3.5 <= average_score < 3.7:
        return "Мастер Откровений (3)"
    elif 3.3 <= average_score < 3.5:
        return "Рыцарь Учебного Пути (3)"
    elif 3.0 <= average_score < 3.3:
        return "Посвящённый Искатель Истины (3)"
    else:
        return "Неофит Знаний"


def save_student_to_db(student_data):
    title_name = assign_title(student_data['average_score'])
    title_id = get_title_id_by_name(title_name)

    if title_id is None:
        raise ValueError(f"Не удалось найти title_id для {title_name}")

    needed_grades = calculate_needed_grades(student_data)

    student_record = StudentRecord(
        student_name=student_data['name'],
        class_name=student_data['class'],
        grades=json.dumps(student_data['subjects'], ensure_ascii=False),
        average_score=round(student_data['average_score'], 2),
        title_id=title_id,
        needed_grades=json.dumps(needed_grades, ensure_ascii=False),
        period=student_data['period']
    )

    db.session.add(student_record)
    print("session added")
    db.session.commit()
    print("session committed")


def get_title_id_by_name(title_name):
    title = Title.query.filter_by(name=title_name).first()
    return title.id if title else None


def process_data(filename):
    print("process data started")
    dataset = get_dataset(filename)
    print("database got")
    processed_students = []

    for student in dataset:
        student_info = {}
        student_info['name'] = student['name']
        print(student_info['name'])
        student_info['class'] = student['class']
        student_info['subjects'] = student.get('subjects', [])
        student_info['overall_average'] = calculate_overall_average(student)
        student_info['average_score'] = student_info['overall_average']

        if 'period' in student:
            student_info['period'] = student['period']
        else:
            print(f"Ошибка: Период отсутствует для студента {student['name']}")

        student_info['needed_grades'] = calculate_needed_grades(student)

        print(student_info)
        print("Приступаем к сохранению в БД")
        save_student_to_db(student_info)
        processed_students.append(student_info)


if __name__ == "__main__":
    process_data("/Users/admin/Desktop/projects/teach.extpankov/web/userfiles/20240824120410.xlsx")
